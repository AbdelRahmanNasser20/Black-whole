"""Fill the Facebook/Meta Commerce catalog product feed from our live inventory.

Takes Meta's official "catalog_products" XLSX template (row 1 = field
instructions, row 2 = field-key header `id,title,description,…`, row 3 = a
sample) and replaces the sample with one row per currently-available lot
(everything `inventory.list_public()` returns — on-hand + won/incoming).

Usage:
    python scripts/build_fb_catalog.py [--template PATH] [--out PATH]

Defaults read `catalog/fb_catalog_template.xlsx` and write
`catalog/fb_catalog_products.xlsx`. Upload the output in Commerce Manager →
Catalog → Data sources → Upload, or point a scheduled feed at it.
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from automation import config  # noqa: F401  (loads .env)
from automation import inventory

BRAND = "Black Whole"
SITE = "https://black-whole.com"
GOOGLE_CATEGORY = "Furniture > Chairs"

# Per-lot descriptive attributes for the optional FB fields. Keyed by lot_id.
# Derived from each lot's title/subtitle/photos.
ATTRS: dict[str, dict[str, str]] = {
    "334":   {"color": "Tan",            "material": "Metal",            "pattern": ""},
    "7126":  {"color": "Red",            "material": "Fabric and metal", "pattern": ""},
    "28505": {"color": "Saffron",        "material": "Vinyl and metal",  "pattern": ""},
    "125":   {"color": "Natural wood",   "material": "Wood",             "pattern": ""},
    "31225": {"color": "Brown",          "material": "Fabric and metal", "pattern": ""},
    "9006":  {"color": "Mauve",          "material": "Fabric and metal", "pattern": "Diamond"},
    "5003":  {"color": "Burgundy",       "material": "Vinyl and metal",  "pattern": ""},
    "2807":  {"color": "Maroon",         "material": "Fabric and metal", "pattern": ""},
    "folder:Orange_Red_Banquet_Chairs_Cypress_242":
             {"color": "Red and orange", "material": "Fabric and metal", "pattern": "Damask"},
    "folder:ATL_Grey_blueish_chairs_399":
             {"color": "Blue",           "material": "Fabric and metal", "pattern": ""},
}

# Statuses we treat as "incoming" (won at auction, pickup pending) vs on-hand.
INCOMING = {"won_pickup"}


def row_for(item: dict) -> dict[str, str]:
    lot_id = item["lot_id"]
    qty = int(item.get("quantity_remaining") or 0)
    price = float(item.get("price_per_chair") or 0)
    attrs = ATTRS.get(lot_id, {})
    status = item.get("status") or ""
    return {
        "id": lot_id,
        "title": (item.get("title") or "").strip()[:200],
        "description": (item.get("description") or "").strip()[:9999],
        "availability": "in stock" if qty > 0 else "out of stock",
        "condition": "used",
        "price": f"{price:.2f} USD",
        "link": f"{SITE}/listings/{lot_id}",
        "image_link": item.get("hero_image_url") or "",
        "brand": BRAND,
        "google_product_category": GOOGLE_CATEGORY,
        "quantity_to_sell_on_facebook": str(qty) if qty > 0 else "",
        "color": attrs.get("color", ""),
        "material": attrs.get("material", ""),
        "pattern": attrs.get("pattern", ""),
        "custom_label_0": "incoming" if status in INCOMING else "on-hand",
        "custom_label_1": item.get("city") or "",
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    here = Path(__file__).resolve().parent.parent
    ap.add_argument("--template", default=str(here / "catalog" / "fb_catalog_template.xlsx"))
    ap.add_argument("--out", default=str(here / "catalog" / "fb_catalog_products.xlsx"))
    args = ap.parse_args()

    wb = load_workbook(args.template)
    ws = wb[wb.sheetnames[0]]

    # Row 2 holds the machine-readable field keys; map key -> column index.
    header = {(c.value or "").strip(): c.column
              for c in ws[2] if c.value}
    missing = {"id", "title", "description", "availability", "condition",
               "price", "link", "image_link", "brand"} - set(header)
    if missing:
        raise SystemExit(f"template missing required headers: {missing}")

    # Clear every data row below the header (row 3+), keeping rows 1-2.
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    items = inventory.list_public()
    no_image = []
    for i, item in enumerate(items):
        r = row_for(item)
        if not r["image_link"]:
            no_image.append(r["id"])
        for key, col in header.items():
            if key in r:
                ws.cell(row=3 + i, column=col, value=r[key])

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {len(items)} products -> {args.out}")
    if no_image:
        print(f"  WARNING: no image_link for: {no_image} (FB requires it)")


if __name__ == "__main__":
    main()
